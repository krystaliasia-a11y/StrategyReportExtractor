"""
Strategy Tester Report Extractor
Reads all .htm files from ../input, extracts key metrics, and writes to ../output/summary.xlsx
"""

import json
import os
import re
import statistics
from pathlib import Path

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


COLUMNS = [
    "Filename",
    "Symbol",
    "Period",
    "Modelling quality",
    "Spread",
    "Initial deposit",
    "Total net profit",
    "Gross profit",
    "Gross loss",
    "Profit factor",
    "Expected payoff",
    "Absolute drawdown",
    "Maximal drawdown",
    "Relative drawdown",
    "Total trades",
    "Short positions",
    "Short positions won %",
    "Long positions",
    "Long positions won %",
    "Profit trades",
    "Profit trades (% of total)",
    "Loss trades",
    "Loss trades (% of total)",
    "Largest profit trade",
    "Largest loss trade",
    "Average profit trade",
    "Average loss trade",
    "Maximum consecutive wins",
    "Maximum consecutive wins (profit in money)",
    "Maximum consecutive losses",
    "Maximum consecutive losses (loss in money)",
    "Maximal consecutive profit",
    "Maximal consecutive profit (count of wins)",
    "Maximal consecutive loss",
    "Maximal consecutive loss (count of losses)",
    "Average consecutive wins",
    "consecutive losses",
    "Parameters",
]

# Excel number formats keyed by column name
_FMT_DECIMAL = "#,##0.00"
_FMT_INT     = "#,##0"
_FMT_PCT     = "0.00%"

COLUMN_FORMATS = {
    "Spread":                                       _FMT_DECIMAL,
    "Initial deposit":                              _FMT_DECIMAL,
    "Total net profit":                             _FMT_DECIMAL,
    "Gross profit":                                 _FMT_DECIMAL,
    "Gross loss":                                   _FMT_DECIMAL,
    "Profit factor":                                _FMT_DECIMAL,
    "Expected payoff":                              _FMT_DECIMAL,
    "Absolute drawdown":                            _FMT_DECIMAL,
    "Maximal drawdown":                             _FMT_DECIMAL,
    "Relative drawdown":                            _FMT_DECIMAL,
    "Total trades":                                 _FMT_INT,
    "Short positions":                              _FMT_INT,
    "Long positions":                               _FMT_INT,
    "Profit trades":                                _FMT_INT,
    "Loss trades":                                  _FMT_INT,
    "Largest profit trade":                         _FMT_DECIMAL,
    "Largest loss trade":                           _FMT_DECIMAL,
    "Average profit trade":                         _FMT_DECIMAL,
    "Average loss trade":                           _FMT_DECIMAL,
    "Maximum consecutive wins":                     _FMT_INT,
    "Maximum consecutive wins (profit in money)":   _FMT_DECIMAL,
    "Maximum consecutive losses":                   _FMT_INT,
    "Maximum consecutive losses (loss in money)":   _FMT_DECIMAL,
    "Maximal consecutive profit":                   _FMT_DECIMAL,
    "Maximal consecutive profit (count of wins)":   _FMT_INT,
    "Maximal consecutive loss":                     _FMT_DECIMAL,
    "Maximal consecutive loss (count of losses)":   _FMT_INT,
    "Average consecutive wins":                     _FMT_INT,
    "consecutive losses":                           _FMT_INT,
    "Modelling quality":                            _FMT_PCT,
    "Short positions won %":                        _FMT_PCT,
    "Long positions won %":                         _FMT_PCT,
    "Profit trades (% of total)":                   _FMT_PCT,
    "Loss trades (% of total)":                     _FMT_PCT,
}


def get_td_text(td):
    return td.get_text(separator=" ", strip=True) if td else ""


def strip_bracket(raw: str) -> str:
    """Remove trailing '(...)' from a string: '464.06 (4.49%)' → '464.06'."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()


def to_float(s) -> "float | str":
    """Convert a string (optionally with commas or %) to float, or '' on failure."""
    if s == "" or s is None:
        return ""
    try:
        return float(str(s).replace(",", "").replace("%", "").strip())
    except ValueError:
        return ""


def to_int(s) -> "int | str":
    """Convert a string to int, or '' on failure."""
    if s == "" or s is None:
        return ""
    try:
        return int(float(str(s).replace(",", "").strip()))
    except ValueError:
        return ""


def to_pct(s) -> "float | str":
    """Convert '45.67%' or '45.67' to 0.4567 for Excel percentage format."""
    if s == "" or s is None:
        return ""
    try:
        return float(str(s).replace("%", "").strip()) / 100
    except ValueError:
        return ""


def parse_modelling_quality(raw: str) -> "float | str":
    """Extract percentage from 'n.nn % (...)' or 'n.nn%' → decimal for Excel."""
    m = re.search(r"([\d.]+)\s*%", raw)
    return float(m.group(1)) / 100 if m else ""


def parse_positions(raw):
    """Parse '35 (100.00%)' into ('35', '100.00%')."""
    m = re.match(r"(\d+)\s*\(([^)]+)\)", raw)
    if m:
        return m.group(1), m.group(2)
    return raw, ""


def parse_period(period_raw: str, filename: str) -> str:
    """Return 'YYYY.MM.DD - YYYY.MM.DD'.
    Start: first date inside the trailing bracket of the Period cell.
    End:   second 8-digit date in the filename (encodes the actual test window).
    """
    start_date = ""
    m = re.search(r"\((\d{4}\.\d{2}\.\d{2})\s*-\s*\d{4}\.\d{2}\.\d{2}\)\s*$", period_raw)
    if m:
        start_date = m.group(1)

    end_date = ""
    fn_m = re.search(r"-(\d{8})-(\d{8})-", filename)
    if fn_m:
        raw_end = fn_m.group(2)
        end_date = f"{raw_end[:4]}.{raw_end[4:6]}.{raw_end[6:8]}"

    if start_date and end_date:
        return f"{start_date} - {end_date}"
    return start_date or end_date or period_raw


def split_num_bracket(raw: str):
    """Parse 'A (B)' into ('A', 'B') for numeric bracket pairs like '32 (530.09)'."""
    m = re.match(r"(-?[\d.]+)\s*\((-?[\d.]+)\)", raw.strip())
    if m:
        return m.group(1), m.group(2)
    return raw, ""


def parse_pip_period_bracket(period_raw: str) -> str:
    """Extract the trailing bracket date range from Period text.
    '1 Hour (H1) ... (2011.01.01 - 2011.12.31)' → '2011.01.01 - 2011.12.31'
    """
    m = re.search(r"\((\d{4}\.\d{2}\.\d{2}\s*-\s*\d{4}\.\d{2}\.\d{2})\)\s*$", period_raw)
    return m.group(1).strip() if m else period_raw


def parse_trades_from_soup(soup, pip_unit: float) -> dict:
    """Parse the trade list table and return per-report pip statistics.

    Pip = abs(open_price - close_price) / pip_unit for each matched trade pair.
    Returns avg, max, min, median; empty string for each when no trades found.
    """
    open_prices = {}  # order_id -> open_price

    in_trade_table = False
    pip_list = []

    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            tds = row.find_all("td")
            if not tds:
                continue

            # Detect the trade table by its grey header row
            if row.get("bgcolor") == "#C0C0C0":
                texts = [td.get_text(strip=True) for td in tds]
                if "Type" in texts and "Order" in texts and "Price" in texts:
                    in_trade_table = True
                    continue

            if not in_trade_table:
                continue
            if len(tds) < 8:
                continue

            trade_type = tds[2].get_text(strip=True).lower()
            order_str  = tds[3].get_text(strip=True)
            price_str  = tds[5].get_text(strip=True)

            try:
                order_id = int(order_str)
                price    = float(price_str)
            except ValueError:
                continue

            if trade_type in ("buy", "sell"):
                open_prices[order_id] = price
            elif trade_type in ("close", "t/p", "s/l"):
                if order_id in open_prices:
                    pip = abs(open_prices[order_id] - price) / pip_unit
                    pip_list.append(round(pip, 2))

    if not pip_list:
        return {"_pip_avg": "", "_pip_max": "", "_pip_min": "", "_pip_median": ""}

    return {
        "_pip_avg":    round(sum(pip_list) / len(pip_list), 2),
        "_pip_max":    max(pip_list),
        "_pip_min":    min(pip_list),
        "_pip_median": round(statistics.median(pip_list), 2),
    }


def parse_report(file_path: str, pip_config: dict = None) -> dict:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()

    soup = BeautifulSoup(content, "html.parser")
    flat = {}  # label -> value collected from the HTML

    for row in soup.find_all("tr"):
        tds = row.find_all("td")
        if not tds:
            continue

        first = tds[0]
        first_colspan = first.get("colspan", "1")
        first_text = get_td_text(first)

        # ── Two-cell header rows: Symbol / Period / Model / Parameters ──────
        # Structure: <td colspan=2>Label</td><td colspan=4>Value</td>
        if first_colspan == "2" and len(tds) == 2:
            second = tds[1]
            second_colspan = second.get("colspan", "1")
            if second_colspan == "4":
                flat[first_text] = get_td_text(second)
            continue

        # ── "Largest" / "Average" rows ───────────────────────────────────────
        # Structure: <td colspan=2 align=right>Largest|Average</td>
        #            <td>profit trade</td><td>value</td>
        #            <td>loss trade</td><td>value</td>
        if first_colspan == "2" and first.get("align") == "right" and len(tds) >= 5:
            context = first_text  # "Largest", "Average", "Maximum", "Maximal", or ""
            if context in ("Largest", "Average", "Maximum", "Maximal"):
                sub1 = get_td_text(tds[1])
                sub2 = get_td_text(tds[3])
                flat[f"{context} {sub1}"] = get_td_text(tds[2])
                flat[f"{context} {sub2}"] = get_td_text(tds[4])
            elif context == "":
                flat[get_td_text(tds[1])] = get_td_text(tds[2])
                flat[get_td_text(tds[3])] = get_td_text(tds[4])
            continue

        # ── Skip rows whose first cell is colspan=2 but don't match above ───
        if first_colspan == "2":
            continue

        # ── Normal rows: pairs of (label, value) across up to 6 cells ───────
        i = 0
        while i + 1 < len(tds):
            label = get_td_text(tds[i])
            value = get_td_text(tds[i + 1])
            if label:
                flat[label] = value
            i += 2

    # ── Map flat dict to the required columns ────────────────────────────────
    result = {"Filename": Path(file_path).name}

    result["Symbol"] = flat.get("Symbol", "")
    result["Period"] = parse_period(flat.get("Period", ""), Path(file_path).name)
    result["Modelling quality"] = parse_modelling_quality(flat.get("Modelling quality", ""))
    result["Spread"]            = to_float(flat.get("Spread", ""))
    result["Initial deposit"]   = to_float(flat.get("Initial deposit", ""))
    result["Total net profit"]  = to_float(flat.get("Total net profit", ""))
    result["Gross profit"]      = to_float(flat.get("Gross profit", ""))
    result["Gross loss"]        = to_float(flat.get("Gross loss", ""))
    result["Profit factor"]     = to_float(flat.get("Profit factor", ""))
    result["Expected payoff"]   = to_float(flat.get("Expected payoff", ""))
    result["Absolute drawdown"] = to_float(flat.get("Absolute drawdown", ""))
    # Strip bracket value (e.g. '464.06 (4.49%)' → 464.06)
    result["Maximal drawdown"]  = to_float(strip_bracket(flat.get("Maximal drawdown", "")))
    # Strip bracket value (e.g. '4.49% (464.06)' → 4.49)
    result["Relative drawdown"] = to_float(strip_bracket(flat.get("Relative drawdown", "")))
    result["Total trades"]      = to_int(flat.get("Total trades", ""))

    short_raw = flat.get("Short positions (won %)", "")
    long_raw  = flat.get("Long positions (won %)", "")
    short_count, short_pct = parse_positions(short_raw)
    long_count,  long_pct  = parse_positions(long_raw)
    result["Short positions"]       = to_int(short_count)
    result["Short positions won %"] = to_pct(short_pct)
    result["Long positions"]        = to_int(long_count)
    result["Long positions won %"]  = to_pct(long_pct)

    profit_trades_raw = flat.get("Profit trades (% of total)", "")
    loss_trades_raw   = flat.get("Loss trades (% of total)", "")
    profit_count, profit_pct = parse_positions(profit_trades_raw)
    loss_count,   loss_pct   = parse_positions(loss_trades_raw)
    result["Profit trades"]              = to_int(profit_count)
    result["Profit trades (% of total)"] = to_pct(profit_pct)
    result["Loss trades"]                = to_int(loss_count)
    result["Loss trades (% of total)"]   = to_pct(loss_pct)

    result["Largest profit trade"] = to_float(flat.get("Largest profit trade", ""))
    result["Largest loss trade"]   = to_float(flat.get("Largest loss trade", ""))
    result["Average profit trade"] = to_float(flat.get("Average profit trade", ""))
    result["Average loss trade"]   = to_float(flat.get("Average loss trade", ""))

    # "Maximum consecutive wins (profit in money)" raw = "32 (530.09)"
    max_cw_raw = flat.get("Maximum consecutive wins (profit in money)", "")
    max_cw_count, max_cw_money = split_num_bracket(max_cw_raw)
    result["Maximum consecutive wins"]                   = to_int(max_cw_count)
    result["Maximum consecutive wins (profit in money)"] = to_float(max_cw_money)

    # "Maximum consecutive losses (loss in money)" raw = "1 (-29.76)"
    max_cl_raw = flat.get("Maximum consecutive losses (loss in money)", "")
    max_cl_count, max_cl_money = split_num_bracket(max_cl_raw)
    result["Maximum consecutive losses"]                  = to_int(max_cl_count)
    result["Maximum consecutive losses (loss in money)"]  = to_float(max_cl_money)

    # "Maximal consecutive profit (count of wins)" raw = "530.09 (32)"
    mxl_cp_raw = flat.get("Maximal consecutive profit (count of wins)", "")
    mxl_cp_profit, mxl_cp_count = split_num_bracket(mxl_cp_raw)
    result["Maximal consecutive profit"]                = to_float(mxl_cp_profit)
    result["Maximal consecutive profit (count of wins)"] = to_int(mxl_cp_count)

    # "Maximal consecutive loss (count of losses)" raw = "-29.76 (1)"
    mxl_cl_raw = flat.get("Maximal consecutive loss (count of losses)", "")
    mxl_cl_loss, mxl_cl_count = split_num_bracket(mxl_cl_raw)
    result["Maximal consecutive loss"]                   = to_float(mxl_cl_loss)
    result["Maximal consecutive loss (count of losses)"] = to_int(mxl_cl_count)

    result["Average consecutive wins"] = to_int(flat.get("Average consecutive wins", ""))
    result["consecutive losses"]       = to_int(flat.get("Average consecutive losses", ""))

    result["Parameters"] = flat.get("Parameters", "")

    # ── Per-trade pip statistics ─────────────────────────────────────────────
    period_raw = flat.get("Period", "")
    result["_period_bracket"] = parse_pip_period_bracket(period_raw)

    if pip_config is not None:
        symbol_raw = flat.get("Symbol", "")
        symbol = symbol_raw.split()[0] if symbol_raw else ""
        pip_unit = pip_config.get(symbol, 0.0001)
        pip_stats = parse_trades_from_soup(soup, pip_unit)
        result.update(pip_stats)
    else:
        result.update({"_pip_avg": "", "_pip_max": "", "_pip_min": "", "_pip_median": ""})

    return result


RED_FILL = PatternFill(fill_type="solid", fgColor="FF0000")

_PARAM_NAMES = [
    "stop_after_earning_N_pct",
    "TP_RATIO_pct",
    "TP_PULLBACK_pct",
    "DROP_RATIO_pct",
    "DROP_BROUNCE_pct",
    "max_orders_per_side",
    "first_max_buy_sell_when_program_starts",
]


def extract_param(params_str: str, param_name: str) -> str:
    """Extract a named parameter value from a parameters string like 'key=value; key2=value2'."""
    m = re.search(rf"{re.escape(param_name)}\s*=\s*([^\s;,]+)", params_str)
    return m.group(1) if m else ""


def add_analysis_sheet(wb, rows: list):
    ws = wb.create_sheet(title="analysis")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Value Name").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=2, value="Value").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = header_align

    first_params = next(
        (r.get("Parameters", "") for r in rows if r.get("Parameters", "")), ""
    )

    real_equity = next(
        (r.get("Initial deposit") for r in rows
         if r.get("Initial deposit") not in ("", None)),
        0,
    ) or 0

    stop_pct = to_float(extract_param(first_params, "stop_after_earning_N_pct"))
    tep = (real_equity * stop_pct / 100) if isinstance(stop_pct, float) and real_equity else None

    # Total net profit aggregates
    tnp_values = [
        r.get("Total net profit") for r in rows
        if r.get("Total net profit") not in ("", None)
    ]
    total_net_profit  = sum(tnp_values)
    count_tnp_gte_tep = sum(1 for v in tnp_values if tep is not None and v >= tep)
    count_tnp_zero    = sum(1 for v in tnp_values if v == 0)
    tnp_neg_values    = [v for v in tnp_values if v < 0]
    count_tnp_neg     = len(tnp_neg_values)
    avg_tnp_neg       = sum(tnp_neg_values) / len(tnp_neg_values) if tnp_neg_values else ""

    # Maximal drawdown aggregates
    mdd_values = [
        r.get("Maximal drawdown") for r in rows
        if r.get("Maximal drawdown") not in ("", None)
    ]
    total_max_drawdown = max(mdd_values) if mdd_values else ""

    def _mdd_count_avg(threshold_pct):
        subset = [v for v in mdd_values if real_equity and v > real_equity * threshold_pct]
        count = len(subset)
        avg   = sum(subset) / len(subset) if subset else ""
        return count, avg

    count_mdd_gt_20pct, avg_mdd_gt_20pct = _mdd_count_avg(0.20)
    count_mdd_gt_30pct, avg_mdd_gt_30pct = _mdd_count_avg(0.30)
    count_mdd_gt_40pct, avg_mdd_gt_40pct = _mdd_count_avg(0.40)
    count_mdd_gt_50pct, avg_mdd_gt_50pct = _mdd_count_avg(0.50)

    analysis_rows = [
        ("Total net profit (sum)",                  total_net_profit),
        ("Count TNP >= TEP",                        count_tnp_gte_tep),
        ("Count TNP = 0",                           count_tnp_zero),
        ("Count TNP < 0",                           count_tnp_neg),
        ("Average TNP < 0",                         avg_tnp_neg),
        ("Maximal drawdown (max)",                  total_max_drawdown),
        ("Count MDD > real_equity * 20%",           count_mdd_gt_20pct),
        ("Count MDD > real_equity * 30%",           count_mdd_gt_30pct),
        ("Count MDD > real_equity * 40%",           count_mdd_gt_40pct),
        ("Count MDD > real_equity * 50%",           count_mdd_gt_50pct),
        ("Average MDD > real_equity * 20%",         avg_mdd_gt_20pct),
        ("Average MDD > real_equity * 30%",         avg_mdd_gt_30pct),
        ("Average MDD > real_equity * 40%",         avg_mdd_gt_40pct),
        ("Average MDD > real_equity * 50%",         avg_mdd_gt_50pct),
    ] + [(name, extract_param(first_params, name)) for name in _PARAM_NAMES]

    for row_idx, (name, value) in enumerate(analysis_rows, start=2):
        name_cell = ws.cell(row=row_idx, column=1, value=name)
        name_cell.alignment = Alignment(vertical="top")
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        val_cell.alignment = Alignment(vertical="top")
        if isinstance(value, float):
            val_cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25


def add_pip_analysis_sheet(wb, rows: list):
    ws = wb.create_sheet(title="pip_analysis")

    headers = ["Period", "Avg Pip", "Max Pip", "Min Pip", "Median Pip"]
    keys    = ["_period_bracket", "_pip_avg", "_pip_max", "_pip_min", "_pip_median"]

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(keys, start=1):
            value           = row.get(key, "")
            cell            = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment  = Alignment(vertical="top")
            if col_idx > 1 and isinstance(value, float):
                cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 14

    ws.freeze_panes = "A2"


def write_excel(rows: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write header
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=False)

            # Apply number / percentage format where defined
            fmt = COLUMN_FORMATS.get(col_name)
            if fmt and value != "":
                cell.number_format = fmt

            # ── Conditional red fill ─────────────────────────────────────────
            # Maximal drawdown: red if Relative drawdown % > 30
            if col_name == "Maximal drawdown":
                rel_dd = row_data.get("Relative drawdown", "")
                try:
                    if float(rel_dd) > 30:
                        cell.fill = RED_FILL
                except (ValueError, TypeError):
                    pass

            # Total net profit: red if negative
            if col_name == "Total net profit" and value != "":
                try:
                    if float(value) < 0:
                        cell.fill = RED_FILL
                except (ValueError, TypeError):
                    pass

    # Auto-size columns (capped at 40; Parameters capped at 60)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, min(len(str(val)), 60))
        cap = 60 if col_name == "Parameters" else 40
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, cap)

    # Freeze header row
    ws.freeze_panes = "A2"

    add_analysis_sheet(wb, rows)
    add_pip_analysis_sheet(wb, rows)

    wb.save(output_path)
    print(f"Saved: {output_path}")


def main():
    base_dir = Path(__file__).parent.parent
    input_dir = base_dir / "input"
    output_dir = base_dir / "output"
    output_dir.mkdir(exist_ok=True)

    pip_config_path = base_dir / "config" / "pip_config.json"
    pip_config = {}
    if pip_config_path.exists():
        with open(pip_config_path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        pip_config = {k: v for k, v in raw.items() if not k.startswith("_")}
    else:
        print(f"Warning: pip_config.json not found at {pip_config_path}, defaulting to 0.0001 per pip")

    htm_files = sorted(input_dir.glob("*.htm")) + sorted(input_dir.glob("*.html"))
    if not htm_files:
        print(f"No .htm / .html files found in {input_dir}")
        return

    rows = []
    for htm_file in htm_files:
        print(f"Processing: {htm_file.name}")
        try:
            row = parse_report(str(htm_file), pip_config)
            rows.append(row)
        except Exception as exc:
            print(f"  ERROR: {exc}")

    output_path = str(output_dir / "summary.xlsx")
    write_excel(rows, output_path)
    print(f"\nDone. {len(rows)} report(s) extracted.")


if __name__ == "__main__":
    main()
